"""
extract_data.py

Reads the three UNIZULU source documents from data/raw/ and turns them into
a flat list of small, labelled text chunks saved to
data/processed/corpus.jsonl (one JSON object per line):

    {"id": "...", "source": "...", "location": "...", "text": "..."}

Run:
    python extract_data.py
"""

import json
import re
from pathlib import Path

import openpyxl
import pdfplumber

RAW_DIR = Path("data/raw")
OUT_DIR = Path("data/processed")
OUT_DIR.mkdir(parents=True, exist_ok=True)

TIMETABLE_FILE = RAW_DIR / "2026_FIRST_SEMESTER_LECTURE_TIME_TABLE_FOR_THE_RICHARDS_BAY_CAMPUS_.xlsx"
CALENDAR_FILE = RAW_DIR / "2026-UNIZULU-General-Calendar.pdf"
HANDBOOK_FILE = RAW_DIR / "Engineering-Department-Handbook-2024_cleaned.pdf"

CHUNK_WORDS = 350   # target words per PDF chunk
CHUNK_OVERLAP = 60  # words of overlap between consecutive chunks


def clean(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


# --------------------------------------------------------------------------
# 1. Timetable (.xlsx) -> one chunk per (day, time slot) with the venues/
#    modules scheduled, using the header rows as column labels.
# --------------------------------------------------------------------------
def extract_timetable(path: Path):
    if not path.exists():
        print(f"[skip] timetable not found: {path}")
        return []

    chunks = []
    wb = openpyxl.load_workbook(path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.__class__.__name__ == "Chartsheet":
            continue  # skip chart-only sheets

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find the header row: the first row containing "DATE" or "TIME SLOT"
        header_idx = None
        for i, row in enumerate(rows[:6]):
            joined = " ".join(str(c) for c in row if c)
            if "TIME SLOT" in joined.upper() or "DATE" in joined.upper():
                header_idx = i
                break
        if header_idx is None:
            header_idx = 0

        headers = rows[header_idx]
        # Forward-fill blank headers (merged cells) with the nearest label to
        # the left, and track building-row labels above (row before header),
        # since these sheets have a two-row header (building, then room).
        building_row = rows[header_idx - 1] if header_idx > 0 else [None] * len(headers)
        filled_building = []
        last = ""
        for v in building_row:
            if v not in (None, "", " "):
                last = str(v).strip()
            filled_building.append(last)

        col_labels = []
        for idx, h in enumerate(headers):
            label = str(h).strip() if h not in (None, "") else f"col{idx}"
            bld = filled_building[idx] if idx < len(filled_building) else ""
            col_labels.append(f"{bld} {label}".strip())

        current_day = ""
        for row in rows[header_idx + 1:]:
            if row is None or all(c in (None, "", " ") for c in row):
                continue
            day_cell = row[0]
            if day_cell not in (None, "", " "):
                current_day = str(day_cell).strip()
            time_slot = row[1] if len(row) > 1 else None
            if time_slot in (None, "", " "):
                continue
            time_slot = str(time_slot).strip()

            entries = []
            for idx in range(2, len(row)):
                val = row[idx]
                if val in (None, "", " "):
                    continue
                label = col_labels[idx] if idx < len(col_labels) else f"col{idx}"
                entries.append(f"{label.strip()} = {str(val).strip()}")

            if not entries:
                continue

            text = (
                f"Sheet: {sheet_name}. Day: {current_day}. Time slot: {time_slot}. "
                f"Scheduled classes/venues: " + "; ".join(entries) + "."
            )
            chunks.append({
                "id": f"timetable::{sheet_name}::{current_day}::{time_slot}",
                "source": "2026 First Semester Lecture Timetable (Richards Bay Campus)",
                "location": f"Sheet '{sheet_name}', {current_day} {time_slot}",
                "text": clean(text),
            })

    print(f"[ok] timetable -> {len(chunks)} chunks")
    return chunks


# --------------------------------------------------------------------------
# 2. Generic PDF chunker (used for both the calendar and the handbook)
# --------------------------------------------------------------------------
def extract_pdf(path: Path, source_label: str):
    if not path.exists():
        print(f"[skip] not found: {path}")
        return []

    chunks = []
    with pdfplumber.open(path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            text = clean(text)
            if not text:
                continue

            words = text.split()
            start = 0
            part = 0
            while start < len(words):
                end = min(start + CHUNK_WORDS, len(words))
                chunk_text = " ".join(words[start:end])
                part += 1
                chunks.append({
                    "id": f"{source_label}::p{page_num}::{part}",
                    "source": source_label,
                    "location": f"Page {page_num}",
                    "text": chunk_text,
                })
                if end == len(words):
                    break
                start = end - CHUNK_OVERLAP

    print(f"[ok] {source_label} -> {len(chunks)} chunks")
    return chunks


def main():
    all_chunks = []
    all_chunks += extract_timetable(TIMETABLE_FILE)
    all_chunks += extract_pdf(CALENDAR_FILE, "2026 UNIZULU General Calendar")
    all_chunks += extract_pdf(HANDBOOK_FILE, "Engineering Department Handbook 2024")

    out_path = OUT_DIR / "corpus.jsonl"
    with open(out_path, "w", encoding="utf-8") as f:
        for c in all_chunks:
            f.write(json.dumps(c, ensure_ascii=False) + "\n")

    print(f"\nWrote {len(all_chunks)} total chunks -> {out_path}")
    if not all_chunks:
        print("WARNING: no chunks produced. Check that files exist in data/raw/ "
              "with the exact expected filenames.")


if __name__ == "__main__":
    main()
